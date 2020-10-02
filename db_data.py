import openpyxl
from guide.models import Account

# command to load the workbook
wb = openpyxl.load_workbook("MCQ.xlsx")

# to keep track of the sheets
count = 1

# iterating over the excel sheets
for page in wb.sheetnames:
    # to work with the particular sheet
    sheet = wb[page]
    if count < 10:
        # print(sheet.title)
        count += 1

        # to iterate over the rows till the end of the sheet
        for row in range(1, sheet.max_row+1):

            # To check whether the first cell is not null
            if sheet.cell(row, column=1).value != None:
                try:

                    # To extract data from a specific cell from the sheet and assigning it to a variable
                    cell = sheet.cell(row, column=2)
                    print(cell.value.lstrip())
                except Exception:
                    pass
