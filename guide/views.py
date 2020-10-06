import openpyxl
from django.shortcuts import render, redirect
from .models import Account, LegalAspectsOfBusiness, CorporateAccounting, PersonnelManagement, Answer
from .filters import LegalAspectsofBusinessFilter, CorporateAccountingFilter, PersonnelManagementFilter


# Create your views here.

def home(request):
    return render(request, 'guide/home.html', {})


def login(request):
    if request.method == 'POST':
        email = request.POST['input-email']
        name = request.POST['input-name']
        Account.objects.create(email=email, name=name)
    return redirect('subjects')


def subjects(request):
    return render(request, "guide/subjects.html")


def legal_aspects_of_business(request):

    questions = LegalAspectsOfBusiness.objects.all()
    answers = Answer.objects.all()

    my_filter = LegalAspectsOfBusinessFilter(request.GET, queryset=questions)
    questions = my_filter.qs

    context = {
        'questions': questions,
        'answers': answers,
        'my_filter': my_filter,
    }
    return render(request, 'guide/legal_aspects_of_business.html', context)


def corporate_accounting(request):
    questions = CorporateAccounting.objects.all()
    answers = Answer.objects.all()

    my_filter = CorporateAccountingFilter(request.GET, queryset=questions)
    questions = my_filter.qs

    context = {
        'questions': questions,
        'answers': answers,
        'my_filter': my_filter,
    }

    return render(request, 'guide/corporate_Accounting.html', context)


def personnel_management(request):
    questions = PersonnelManagement.objects.all()
    answers = Answer.objects.all()

    my_filter = PersonnelManagementFilter(request.GET, queryset=questions)
    questions = my_filter.qs

    context = {
        'questions': questions,
        'answers': answers,
        'my_filter': my_filter,
    }

    return render(request, 'guide/personnel_management.html', context)


# fuction to extract data from spreadsheet and upload in DB
def corporate_accounting_questions(request):
    # command to load the workbook
    work_book = openpyxl.load_workbook("MCQ.xlsx")

    # to keep track of the sheets

    # iterating over the excel sheets
    for page in work_book.sheetnames:
        # to work with the particular sheet
        sheet = work_book[page]
        if sheet.title == "Page 27" or sheet.title == "Page 28" or sheet.title == "Page 29" or sheet.title == "Page 30" or sheet.title == "Page 31" or sheet.title == "Page 32" or sheet.title == "Page 33" or sheet.title == "Page 34" or sheet.title == "Page 35":
            print(sheet.title)

            # to iterate over the rows till the end of the sheet
            for row in range(1, sheet.max_row+1):

                # To check whether the first cell is not null
                if sheet.cell(row, column=1).value != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        cell2 = sheet.cell(row, column=2)
                        s_no = cell1.value
                        question = cell2.value.lstrip().lower()
                        print(s_no)
                        print(question)

                        CorporateAccounting.objects.create(
                            s_no=s_no, question=question)

                    except Exception:
                        pass
    return render(request, "guide/corporate_accounting_questions.html")


def corporate_accounting_answers(request):
    # command to load the workbook
    work_book = openpyxl.load_workbook("MCQ.xlsx")

    # to keep track of the sheets

    # iterating over the excel sheets
    for page in work_book.sheetnames:
        # to work with the particular sheet
        sheet = work_book[page]
        if sheet.title == "Page 35" or sheet.title == "Page 36" or sheet.title == "Page 37" or sheet.title == "Page 37":
            print(sheet.title)

            # to iterate over the rows till the end of the sheet
            for row in range(1, sheet.max_row+1):

                # To check whether the first cell is not null
                if sheet.cell(row, column=1).value != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        # cell2 = sheet.cell(row, column=2)
                        s_no = cell1.value
                        answer = s_no[4:]
                        s_no = s_no[0:2]
                        print(s_no)
                        print(s_no[0:2])
                        print(answer)

                        Answer.objects.create(s_no=s_no, answer=answer)
                        question = CorporateAccounting.objects.get(s_no=s_no)
                        question.answer = answer
                        question.save()

                    except Exception:
                        pass
    return render(request, "guide/corporate_accounting_answers.html")


# fuction to extract data from spreadsheet and upload in DB
def legal_aspects_questions(request):
    # command to load the workbook
    work_book = openpyxl.load_workbook("MCQ.xlsx")

    # to keep track of the sheets
    count = 1

    # iterating over the excel sheets
    for page in work_book.sheetnames:
        # to work with the particular sheet
        sheet = work_book[page]
        if count < 11:
            print(sheet.title)
            count += 1

            # to iterate over the rows till the end of the sheet
            for row in range(1, sheet.max_row+1):

                # To check whether the first cell is not null
                if sheet.cell(row, column=1).value != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        cell2 = sheet.cell(row, column=2)
                        s_no = cell1.value
                        question = cell2.value.lstrip().lower()
                        print(question)
                        print(s_no)
                        Question.objects.create(s_no=s_no, question=question)

                    except Exception:
                        pass
    return render(request, "guide/legal_aspects_questions.html")


def legal_aspects_answers(request):
    # command to load the workbook
    work_book = openpyxl.load_workbook("MCQ.xlsx")

    # to keep track of the sheets

    # iterating over the excel sheets
    for page in work_book.sheetnames:
        # to work with the particular sheet
        sheet = work_book[page]
        if sheet.title == "Page 10" or sheet.title == "Page 11" or sheet.title == "Page 12" or sheet.title == "Page 13" or sheet.title == "Page 14":
            print(sheet.title)

            # to iterate over the rows till the end of the sheet
            for row in range(1, sheet.max_row+1):

                # To check whether the first cell is not null
                if sheet.cell(row, column=1).value != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        cell2 = sheet.cell(row, column=2)
                        s_no = cell1.value
                        answer = cell2.value.strip().lower()
                        print(answer)
                        print(s_no)
                        Answer.objects.create(s_no=s_no, answer=answer)
                        question = Question.objects.get(s_no=s_no)
                        question.answer = answer
                        question.save()

                    except Exception:
                        pass
    return render(request, "guide/legal_aspects_answers.html")


def personnel_management_questions(request):
    # command to load the workbook
    work_book = openpyxl.load_workbook("MCQ.xlsx")

    # to keep track of the sheets

    # iterating over the excel sheets
    for page in work_book.sheetnames:
        # to work with the particular sheet
        sheet = work_book[page]
        if sheet.title == "Page 48":
            # or sheet.title == "Page 49" or sheet.title == "Page 50" or sheet.title == "Page 51" or sheet.title == "Page 52" or sheet.title == "Page 53" or sheet.title == "Page 54" or sheet.title == "Page 55" or sheet.title == "Page 56" or sheet.title == "Page 57" or sheet.title == "Page 58" or sheet.title == "Page 59":
            print(sheet.title)

            # to iterate over the rows till the end of the sheet
            for row in range(1, sheet.max_row+1):

                # To check whether the first cell is not null
                if sheet.cell(row, column=1).value != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        # cell2 = sheet.cell(row, column=2)
                        # s_no = cell1.value
                        # question = cell2.value.lstrip().lower()
                        # print(s_no)
                        # print(question)
                        s_no = cell1.value
                        if s_no[0] == "1" or s_no[0] == "2" or s_no[0] == "3" or s_no[0] == "4" or s_no[0] == "5" or s_no[0] == "6" or s_no[0] == "7" or s_no[0] == "8" or s_no[0] == "9":
                            question = s_no[3:].lstrip().lower()
                            s_no = s_no[0:1]
                            print(s_no)
                            print(question)

                            PersonnelManagement.objects.create(
                                s_no=s_no, question=question)
                    except Exception:
                        pass
    return render(request, "guide/personnel_management_questions.html")


def personnel_management_answers(request):
    # command to load the workbook
    work_book = openpyxl.load_workbook("MCQ.xlsx")

    # to keep track of the sheets

    # iterating over the excel sheets
    for page in work_book.sheetnames:
        # to work with the particular sheet
        sheet = work_book[page]
        if sheet.title == "Page 60" or sheet.title == "Page 61" or sheet.title == "Page 62":
            # or sheet.title == "Page 51":
            print(sheet.title)

         # to iterate over the rows till the end of the sheet
            for row in range(1, sheet.max_row+1):

                # To check whether the first cell is not null
                if sheet.cell(row, column=1).value != None:
                    try:
                        cell1 = sheet.cell(row, column=1)
                        # cell2 = sheet.cell(row, column=2)
                        s_no = cell1.value
                        answer = s_no[4:]
                        s_no = s_no[0:2]
                        print(s_no)
                        print(s_no[0:2])
                        print(answer)

                        Answer.objects.create(s_no=s_no, answer=answer)
                        question = PersonnelManagement.objects.get(s_no=s_no)
                        question.answer = answer
                        question.save()

                    except Exception:
                        pass
    return render(request, "guide/personnel_management_answers.html")
